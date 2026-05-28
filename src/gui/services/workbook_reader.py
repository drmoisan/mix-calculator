"""Workbook tab discovery and preview reader (Excel read boundary).

This module provides the GUI's Excel inspection seam used by the source-selection
presenter: enumerate the worksheet tabs in a workbook and read a bounded,
string-rendered preview of a chosen tab. It uses ``openpyxl.load_workbook`` in
read-only mode and reads cell values in-process; it creates no temporary files
and performs no transform logic.

Responsibilities:
    - ``get_sheet_names`` lists the worksheet titles in a workbook.
    - ``read_sheet_preview`` reads at most ``max_rows`` rows of one sheet and
      returns their cell values as strings for display.

Boundaries:
    - The only I/O is the openpyxl read of the workbook path/buffer. The reader
      holds no pipeline or transform logic.

Usage:
    Constructed at the composition root and injected into the source-selection
    presenter behind :class:`WorkbookReaderProtocol`; tests substitute a fake.
"""

from __future__ import annotations

from typing import IO, Protocol

from openpyxl import load_workbook

__all__ = [
    "WorkbookReader",
    "WorkbookReaderProtocol",
]

# Default cap on preview rows to bound rendering cost for very large tabs.
_DEFAULT_MAX_ROWS = 200

# Accepted workbook sources: a filesystem path or a binary file-like buffer
# (the in-memory tests pass a BytesIO).
WorkbookSource = str | IO[bytes]


class WorkbookReaderProtocol(Protocol):
    """Contract for the workbook tab/preview reader.

    Purpose:
        Decouple the source-selection presenter from openpyxl so tests can
        inject a fake reader.

    Responsibilities:
        Enumerate worksheet tab names and read a bounded preview of one tab.
        Implementations perform the Excel read; the Protocol carries only the
        call surface the presenter depends on.
    """

    def get_sheet_names(self, path: str) -> list[str]:
        """Return the worksheet-tab names in the workbook at ``path``.

        Args:
            path: The workbook path to inspect.

        Returns:
            The worksheet titles in workbook order.
        """
        ...

    def read_sheet_preview(
        self, path: str, sheet_name: str, max_rows: int = _DEFAULT_MAX_ROWS
    ) -> list[list[str]]:
        """Return a bounded string preview of one worksheet.

        Args:
            path: The workbook path to read.
            sheet_name: The worksheet to preview.
            max_rows: Maximum number of rows to return.

        Returns:
            The preview rows, each a list of string cell values.
        """
        ...


class WorkbookReader:
    """openpyxl-backed reader for worksheet tabs and previews.

    Purpose:
        Implement :class:`WorkbookReaderProtocol` against ``openpyxl`` for the
        GUI's tab discovery and preview.

    Responsibilities:
        List sheet names and read a bounded, string-rendered preview of a chosen
        sheet. It does not modify the workbook and writes no files.

    Usage:
        Accepts a path (the production case) or a binary buffer (tests). The
        ``read_only=True`` load keeps memory bounded for large workbooks.

    Key invariants:
        ``read_sheet_preview`` never returns more than ``max_rows`` rows. Cell
        values are rendered as strings; a blank cell (``None``) becomes ``""``.
    """

    def get_sheet_names(self, path: WorkbookSource) -> list[str]:
        """Return the worksheet-tab names in the workbook.

        Args:
            path: Filesystem path or binary buffer of the workbook.

        Returns:
            The worksheet titles in workbook order.

        Raises:
            Exception: Propagated from ``openpyxl.load_workbook`` when the
                workbook cannot be opened (for example an invalid/corrupt file).

        Side effects:
            Opens and closes the workbook in read-only mode.
        """
        workbook = load_workbook(path, read_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def read_sheet_preview(
        self,
        path: WorkbookSource,
        sheet_name: str,
        max_rows: int = _DEFAULT_MAX_ROWS,
    ) -> list[list[str]]:
        """Read a bounded, string-rendered preview of one worksheet.

        Args:
            path: Filesystem path or binary buffer of the workbook.
            sheet_name: The worksheet to preview.
            max_rows: Maximum number of rows to return (default 200).

        Returns:
            At most ``max_rows`` rows, each a list of string cell values. A blank
            cell is rendered as the empty string.

        Raises:
            Exception: Propagated from ``openpyxl.load_workbook`` when the
                workbook cannot be opened.
            KeyError: When ``sheet_name`` is not present in the workbook.

        Side effects:
            Opens and closes the workbook in read-only mode.
        """
        workbook = load_workbook(path, read_only=True)
        try:
            worksheet = workbook[sheet_name]
            rows: list[list[str]] = []
            # Walk the worksheet rows up to the max_rows cap so a very large tab
            # never materializes more than the bounded preview. Each cell value
            # is rendered as a string, with a blank cell becoming "".
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index >= max_rows:
                    break
                rows.append(["" if cell is None else str(cell) for cell in row])
            return rows
        finally:
            workbook.close()
