"""Unit tests for :mod:`src.gui.pipeline_service`.

Covers the import seam (``import_le``/``import_aop``/``import_skulu`` and
``import_sources`` with the SKU_LU default), the ``run_pipeline`` orchestration
over fabricated imported frames, and the negative flow where a loader
``ValueError`` propagates unchanged. All Excel inputs are in-memory ``BytesIO``
workbooks; no temporary files are created. The LE and AOP workbooks reuse the
shared builders from ``tests.le_fixtures`` and ``tests.aop_fixtures``; the SKU_LU
workbook and the combined LE+AOP+SKU_LU single-workbook fixture are fabricated
in-file because no shared SKU_LU builder exists in the suite. Fabricated data
only; no confidential ``SKU Description`` or ``Category`` values appear.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import Workbook

from src.gui.pipeline_service import ImportSpec, PipelineService
from src.gui.services.db_service import DbService
from tests.aop_fixtures import aop_header_without_key, make_aop_row
from tests.le_fixtures import make_row, patch_connect, source_header_without_key

if TYPE_CHECKING:
    import pytest

# Fabricated GtN labels (LE) and AOP Type rows shared across both sheets so the
# pipeline produces normal, tie-able lines. The Type segment is distinct per
# label so the rebuilt KEY differs per label and rows do not collapse.
_GTN_LABELS = ["Gross Sales", "Lbs", "Off Invoice", "Trade", "Non-Trade"]
_AOP_TYPES = [
    ("Gross Sales", 100.0),
    ("LBs", 10.0),
    ("Off Invoice $", 5.0),
    ("Trade Spend $", 2.0),
    ("Non-Trade $", 1.0),
]
_SKUS = ("SKU-001", "SKU-002")

# The expected derived-table keys produced by run_pipeline (intermediate
# pivots/lookups plus the run_transforms output), excluding the import frames.
_EXPECTED_DERIVED_KEYS = {
    "le_wide",
    "aop_wide",
    "customer_lu",
    "aop_norm",
    "le_norm",
    "aop_vs_le",
    "mix_base",
    "rate_impacts",
    "mix_rollup_1",
    "mix_1_sku",
    "mix_rollup_2",
    "mix_2_category",
    "mix_rollup_3",
    "mix_3_customer",
    "mix_rollup_4",
    "mix_4_country",
    "mix_0_detail",
    "q1_results_by_sku",
}


def _le_rows() -> list[dict[str, object]]:
    """Build fabricated LE source rows: one GtN label row per SKU."""
    rows: list[dict[str, object]] = []
    # Two SKUs in one customer; each GtN label carries a May..Dec-weighted month
    # profile so the loader's YTG = sum(May..Dec) derivation is exercised.
    for sku in _SKUS:
        for label in _GTN_LABELS:
            value = 12.0 if label == "Lbs" else 100.0
            months = [0.0, 0.0, 0.0, 0.0, *([value / 8.0] * 8)]
            rows.append(
                make_row(
                    customer="Acme Foods",
                    sku=sku,
                    type_=label,
                    ppg="PPG-1",
                    months=months,
                    gtn=label,
                    description="Widget",
                )
            )
    return rows


def _aop_rows() -> list[dict[str, object]]:
    """Build fabricated AOP source rows: one Type row per SKU."""
    rows: list[dict[str, object]] = []
    # Mirror the LE SKUs/customer so the AOP-vs-LE comparison ties out.
    for sku in _SKUS:
        for type_label, ytg in _AOP_TYPES:
            months = [0.0, 0.0, 0.0, 0.0, *([ytg / 8.0] * 8)]
            rows.append(
                make_aop_row(
                    customer="Acme Foods",
                    sku=sku,
                    type_=type_label,
                    months=months,
                    customer_master="Master Group",
                    description="Widget",
                    ppg="PPG-1",
                )
            )
    return rows


def _build_le_workbook() -> io.BytesIO:
    """Build a standalone LE workbook (header on Excel row 3)."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "LE-8 + 4"
    header = source_header_without_key()
    sheet.append(["leading note row 1"])
    sheet.append(["leading note row 2"])
    sheet.append(header)
    # Emit each LE row in the header's column order.
    for record in _le_rows():
        sheet.append([record.get(column) for column in header])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_aop_workbook() -> io.BytesIO:
    """Build a standalone AOP workbook (header on Excel row 3)."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "AOP1"
    header = aop_header_without_key()
    sheet.append(["leading note row 1"])
    sheet.append(["leading note row 2"])
    sheet.append(header)
    for record in _aop_rows():
        sheet.append([record.get(column) for column in header])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _append_skulu_sheet(workbook: Workbook) -> None:
    """Append a fabricated SKU_LU sheet (header on Excel row 1).

    The header is on Excel row 1 with columns SKU, SKU Description, Category,
    International. Values are fabricated only; no confidential data appears.
    """
    sheet = workbook.create_sheet("SKU_LU")
    sheet.append(["SKU", "SKU Description", "Category", "International"])
    sheet.append(["SKU-001", "Widget A", "Category X", 0])
    sheet.append(["SKU-002", "Widget B", "Category Y", 1])


def _build_skulu_workbook() -> io.BytesIO:
    """Build a standalone SKU_LU workbook (header on Excel row 1)."""
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)
    _append_skulu_sheet(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _append_le_sheet(workbook: Workbook) -> None:
    """Append the LE-8 + 4 sheet (header on Excel row 3) to a workbook."""
    sheet = workbook.create_sheet("LE-8 + 4")
    header = source_header_without_key()
    sheet.append(["leading note row 1"])
    sheet.append(["leading note row 2"])
    sheet.append(header)
    for record in _le_rows():
        sheet.append([record.get(column) for column in header])


def _append_aop_sheet(workbook: Workbook) -> None:
    """Append the AOP1 sheet (header on Excel row 3) to a workbook."""
    sheet = workbook.create_sheet("AOP1")
    header = aop_header_without_key()
    sheet.append(["leading note row 1"])
    sheet.append(["leading note row 2"])
    sheet.append(header)
    for record in _aop_rows():
        sheet.append([record.get(column) for column in header])


def _build_combined_workbook() -> io.BytesIO:
    """Build one in-memory workbook carrying LE, AOP, and SKU_LU sheets.

    Used to exercise the ``import_sources`` SKU_LU default, where the SKU_LU
    workbook defaults to the LE/AOP workbook when ``skulu_path`` is empty.
    """
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)
    _append_le_sheet(workbook)
    _append_aop_sheet(workbook)
    _append_skulu_sheet(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _patch_loaders(monkeypatch: pytest.MonkeyPatch, buffer: io.BytesIO) -> None:
    """Patch the three loader reads to read a single shared in-memory buffer."""
    from src import load_aop, load_skulu, normalize_le

    real_load_source = normalize_le.load_source
    real_load_aop = load_aop.load_aop
    real_load_skulu = load_skulu.load_skulu

    def _fake_load_source(
        _path: str, sheet: str, *, key_mismatch: str = "prompt"
    ) -> object:
        buffer.seek(0)
        return real_load_source(buffer, sheet, key_mismatch=key_mismatch)

    def _fake_load_aop(
        _path: str, *, sheet: str = "AOP1", key_mismatch: str = "prompt"
    ) -> object:
        buffer.seek(0)
        return real_load_aop(buffer, sheet=sheet, key_mismatch=key_mismatch)

    def _fake_load_skulu(_path: str, *, sheet: str = "SKU_LU") -> object:
        buffer.seek(0)
        return real_load_skulu(buffer, sheet=sheet)

    monkeypatch.setattr("src.normalize_le.load_source", _fake_load_source)
    monkeypatch.setattr("src.load_aop.load_aop", _fake_load_aop)
    monkeypatch.setattr("src.load_skulu.load_skulu", _fake_load_skulu)


def test_import_le_returns_normalized_frame(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """import_le returns a normalized LE frame with the KEY column."""
    # Arrange
    _patch_loaders(monkeypatch, _build_le_workbook())
    service = PipelineService()

    # Act
    frame = service.import_le("ignored.xlsx", "LE-8 + 4")

    # Assert: the normalized output carries a KEY and one row per unique KEY.
    assert "KEY" in frame.columns
    assert len(frame) == len(_SKUS) * len(_GTN_LABELS)


def test_import_aop_returns_validated_frame(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """import_aop returns a validated AOP frame with the KEY column."""
    # Arrange
    _patch_loaders(monkeypatch, _build_aop_workbook())
    service = PipelineService()

    # Act
    frame = service.import_aop("ignored.xlsx", "AOP1")

    # Assert
    assert "KEY" in frame.columns
    assert len(frame) == len(_SKUS) * len(_AOP_TYPES)


def test_import_skulu_returns_lookup_frame(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """import_skulu returns the cleaned SKU lookup with Country mapped."""
    # Arrange
    _patch_loaders(monkeypatch, _build_skulu_workbook())
    service = PipelineService()

    # Act
    frame = service.import_skulu("ignored.xlsx", "SKU_LU")

    # Assert: International is renamed to Country and codes map to labels.
    assert "Country" in frame.columns
    assert set(frame["Country"]) == {"US", "Canada"}


def test_import_sources_returns_all_three_keys(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """import_sources returns the LE/aop/sku_lu frames from a spec."""
    # Arrange: a combined workbook so all three sheets resolve from one buffer.
    _patch_loaders(monkeypatch, _build_combined_workbook())
    service = PipelineService()
    spec = ImportSpec(
        le_path="combined.xlsx",
        le_sheet="LE-8 + 4",
        aop_path="combined.xlsx",
        aop_sheet="AOP1",
        skulu_path="combined.xlsx",
        skulu_sheet="SKU_LU",
    )

    # Act
    tables = service.import_sources(spec)

    # Assert
    assert set(tables) == {"LE", "aop", "sku_lu"}


def test_import_sources_defaults_skulu_to_le_workbook(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """An empty skulu_path defaults the SKU_LU workbook to the LE workbook."""
    # Arrange: the combined workbook exercises the default because the SKU_LU
    # sheet lives on the same workbook as LE and AOP.
    paths_read: list[str] = []
    from src import load_skulu

    real_load_skulu = load_skulu.load_skulu
    buffer = _build_combined_workbook()
    _patch_loaders(monkeypatch, buffer)

    def _recording_load_skulu(path: str, *, sheet: str = "SKU_LU") -> object:
        # Record the resolved SKU_LU path so the default can be asserted.
        paths_read.append(path)
        buffer.seek(0)
        return real_load_skulu(buffer, sheet=sheet)

    monkeypatch.setattr("src.load_skulu.load_skulu", _recording_load_skulu)
    service = PipelineService()
    spec = ImportSpec(
        le_path="le-workbook.xlsx",
        le_sheet="LE-8 + 4",
        aop_path="aop-workbook.xlsx",
        aop_sheet="AOP1",
        skulu_path="",
        skulu_sheet="SKU_LU",
    )

    # Act
    service.import_sources(spec)

    # Assert: the SKU_LU loader received the LE workbook path (the default).
    assert paths_read == ["le-workbook.xlsx"]


def test_import_le_propagates_loader_value_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A loader ValueError propagates unchanged from import_le."""
    # Arrange: patch the LE loader to raise the kind of ValueError loaders emit.
    import pytest as _pytest

    def _raise(_path: str, _sheet: str, *, key_mismatch: str = "prompt") -> object:
        raise ValueError("Source schema mismatch: could not resolve column")

    monkeypatch.setattr("src.normalize_le.load_source", _raise)
    service = PipelineService()

    # Act / Assert: the same ValueError surfaces to the caller unchanged.
    with _pytest.raises(ValueError, match="could not resolve column"):
        service.import_le("ignored.xlsx", "LE-8 + 4")


def test_run_pipeline_returns_full_derived_key_set(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """run_pipeline produces every derived table without raising."""
    # Arrange: import all three frames from the combined workbook, then run.
    _patch_loaders(monkeypatch, _build_combined_workbook())
    service = PipelineService()
    spec = ImportSpec(
        le_path="combined.xlsx",
        le_sheet="LE-8 + 4",
        aop_path="combined.xlsx",
        aop_sheet="AOP1",
        skulu_path="combined.xlsx",
        skulu_sheet="SKU_LU",
    )
    tables = service.import_sources(spec)

    # Act
    derived = service.run_pipeline(tables)

    # Assert: the full derived-table key set is present.
    assert _EXPECTED_DERIVED_KEYS.issubset(set(derived))


def test_save_to_db_delegates_to_db_service(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """save_to_db persists the working tables via the injected DbService."""
    # Arrange: an in-memory connection and a service over a real DbService.
    con = patch_connect(monkeypatch)
    service = PipelineService(db_service=DbService())
    tables = {"LE": pd.DataFrame({"KEY": ["k1"], "FY": [10.0]})}

    # Act
    service.save_to_db(tables, "ignored.db")
    loaded = service.open_db("ignored.db")

    # Assert: the table survives the save/open round-trip via the seam.
    assert set(loaded) == {"LE"}
    pd.testing.assert_frame_equal(loaded["LE"], tables["LE"])
    con.real_close()


def test_open_db_returns_loaded_tables(monkeypatch: pytest.MonkeyPatch) -> None:
    """open_db loads the tables previously saved through the service."""
    # Arrange
    con = patch_connect(monkeypatch)
    service = PipelineService()
    service.save_to_db({"mix_rollup_4": pd.DataFrame({"value": [42.0]})}, "x.db")

    # Act
    loaded = service.open_db("x.db")

    # Assert
    assert list(loaded["mix_rollup_4"]["value"]) == [42.0]
    con.real_close()
