"""Unit tests for :mod:`src.gui.services.workbook_reader`.

Covers ``WorkbookReader.get_sheet_names`` (single-tab, multi-tab, duplicate tab
names) and ``read_sheet_preview`` (string conversion, the ``max_rows`` bound on
a large fabricated sheet, blank-cell handling), plus a negative flow for an
unreadable/invalid workbook. All workbooks are in-memory ``BytesIO``; no
temporary files are created. Fabricated data only; no confidential values.
"""

from __future__ import annotations

import io
import zipfile

import pytest
from openpyxl import Workbook

from src.gui.services.workbook_reader import WorkbookReader


def _build_workbook(sheet_names: list[str]) -> io.BytesIO:
    """Build an in-memory workbook with the given sheet titles.

    Args:
        sheet_names: The worksheet titles to create, in order.

    Returns:
        A ``BytesIO`` positioned at offset 0 containing the workbook bytes.
    """
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)
    # Create one sheet per requested title so the reader enumerates them in order.
    for name in sheet_names:
        workbook.create_sheet(name)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_workbook_with_rows(rows: list[list[object]]) -> io.BytesIO:
    """Build a single-sheet workbook whose first sheet carries ``rows``.

    Args:
        rows: The cell rows to append to the sheet.

    Returns:
        A ``BytesIO`` positioned at offset 0 containing the workbook bytes.
    """
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data"
    # Append each row as-is so the preview reader sees the fabricated cells.
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def test_get_sheet_names_single_tab() -> None:
    """get_sheet_names returns the single tab of a one-sheet workbook."""
    # Arrange
    reader = WorkbookReader()
    workbook = _build_workbook(["OnlyTab"])

    # Act
    names = reader.get_sheet_names(workbook)

    # Assert
    assert names == ["OnlyTab"]


def test_get_sheet_names_multi_tab() -> None:
    """get_sheet_names returns every tab in workbook order."""
    # Arrange
    reader = WorkbookReader()
    workbook = _build_workbook(["AOP1", "LE-8 + 4", "SKU_LU"])

    # Act
    names = reader.get_sheet_names(workbook)

    # Assert
    assert names == ["AOP1", "LE-8 + 4", "SKU_LU"]


def test_get_sheet_names_duplicate_titles_are_disambiguated() -> None:
    """openpyxl disambiguates duplicate tab titles; all tabs are returned."""
    # Arrange: openpyxl appends a suffix to a duplicate title, so the workbook
    # still has two distinct sheets that the reader must both enumerate.
    reader = WorkbookReader()
    workbook = _build_workbook(["Tab", "Tab"])

    # Act
    names = reader.get_sheet_names(workbook)

    # Assert: two tabs are present and the first keeps the requested title.
    assert len(names) == 2
    assert names[0] == "Tab"


def test_read_sheet_preview_converts_cells_to_strings() -> None:
    """read_sheet_preview renders mixed cell types as strings."""
    # Arrange: a row mixing text, int, and float values.
    reader = WorkbookReader()
    workbook = _build_workbook_with_rows([["Acme Foods", 1, 2.5]])

    # Act
    preview = reader.read_sheet_preview(workbook, "Data")

    # Assert: every cell is a string.
    assert preview == [["Acme Foods", "1", "2.5"]]


def test_read_sheet_preview_bounds_rows_by_max_rows() -> None:
    """read_sheet_preview returns at most max_rows rows for a large sheet."""
    # Arrange: a sheet with more rows than the requested cap.
    reader = WorkbookReader()
    rows: list[list[object]] = [[f"row-{i}"] for i in range(500)]
    workbook = _build_workbook_with_rows(rows)

    # Act
    preview = reader.read_sheet_preview(workbook, "Data", max_rows=10)

    # Assert
    assert len(preview) == 10
    assert preview[0] == ["row-0"]


def test_read_sheet_preview_renders_blank_cells_as_empty_string() -> None:
    """A blank (None) cell is rendered as the empty string."""
    # Arrange: a row with a gap (None) between two values.
    reader = WorkbookReader()
    workbook = _build_workbook_with_rows([["a", None, "c"]])

    # Act
    preview = reader.read_sheet_preview(workbook, "Data")

    # Assert
    assert preview == [["a", "", "c"]]


def test_get_sheet_names_invalid_workbook_raises() -> None:
    """An unreadable/invalid workbook surfaces an error."""
    # Arrange: bytes that are not a valid .xlsx workbook.
    reader = WorkbookReader()
    invalid = io.BytesIO(b"this is not a workbook")

    # Act / Assert: an .xlsx is a zip container, so non-zip bytes fail at the
    # zip layer with a specific BadZipFile error surfaced to the caller.
    with pytest.raises(zipfile.BadZipFile):
        reader.get_sheet_names(invalid)
