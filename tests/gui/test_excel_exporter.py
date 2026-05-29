"""Unit tests for :mod:`src.gui.exporters.excel_exporter`.

Verifies ``ExcelExporter.export`` writes one worksheet per selected table by
exporting to an in-memory ``BytesIO`` target and reading the sheets back via
pandas/openpyxl (frame round-trip). Covers a subset selection and the
``format_name`` value. No temp files; fabricated data only.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook

from src.gui.exporters.excel_exporter import ExcelExporter
from src.pandas_io import read_excel_sheet


def _fabricated_tables() -> dict[str, pd.DataFrame]:
    """Return a fabricated three-table set for export tests."""
    return {
        "le_wide": pd.DataFrame({"KEY": ["k1", "k2"], "FY": [10.0, 20.0]}),
        "aop_wide": pd.DataFrame({"KEY": ["k3"], "FY": [30.0]}),
        "mix_rollup_4": pd.DataFrame({"value": [42.0]}),
    }


def test_format_name_is_excel() -> None:
    """The exporter reports the Excel format name."""
    # Arrange / Act / Assert
    assert ExcelExporter().format_name == "Excel"


def test_export_writes_one_sheet_per_selected_table() -> None:
    """export writes a worksheet per selected name, round-tripping content."""
    # Arrange: export a subset (two of three tables) to an in-memory buffer.
    exporter = ExcelExporter()
    tables = _fabricated_tables()
    target = io.BytesIO()

    # Act
    exporter.export(tables, ["le_wide", "mix_rollup_4"], target)

    # Assert: exactly the selected sheets exist (enumerated via openpyxl) and
    # each sheet's content round-trips (read through the typed pandas boundary).
    target.seek(0)
    workbook = load_workbook(target, read_only=True)
    try:
        assert set(workbook.sheetnames) == {"le_wide", "mix_rollup_4"}
    finally:
        workbook.close()
    target.seek(0)
    le_back = read_excel_sheet(target, sheet_name="le_wide", header=0)
    # Excel does not preserve the int-vs-float distinction for whole numbers, so
    # compare values without enforcing dtype across the cross-format round-trip.
    pd.testing.assert_frame_equal(le_back, tables["le_wide"], check_dtype=False)
    target.seek(0)
    rollup_back = read_excel_sheet(target, sheet_name="mix_rollup_4", header=0)
    pd.testing.assert_frame_equal(
        rollup_back, tables["mix_rollup_4"], check_dtype=False
    )


def test_export_subset_excludes_unselected_tables() -> None:
    """Unselected tables are not written to the workbook."""
    # Arrange
    exporter = ExcelExporter()
    tables = _fabricated_tables()
    target = io.BytesIO()

    # Act: select only one table.
    exporter.export(tables, ["aop_wide"], target)

    # Assert: the unselected tables are absent (enumerate sheets via openpyxl).
    target.seek(0)
    workbook = load_workbook(target, read_only=True)
    try:
        assert set(workbook.sheetnames) == {"aop_wide"}
    finally:
        workbook.close()
