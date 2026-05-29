"""End-to-end integration test for the mix-pipeline GUI seam.

Drives the full scenario through the real :class:`PipelineService`,
:class:`DbService`, :class:`ExcelExporter`, :class:`CsvExporter`, and the
presenters; fakes only at the Qt-view boundary. The scenario is:

    select -> import (LE/AOP/SKU_LU via BytesIO workbooks) -> run -> save (to
    ``sqlite3.connect(":memory:")``) -> reopen -> export (Excel to BytesIO, CSV
    via an in-memory write seam).

All Excel inputs are in-memory ``BytesIO``; the SQLite database is in-memory via
the ``PersistentConnection`` / ``patch_connect`` pattern; CSV is captured via an
injected ``StringIO``-backed write seam. No temp files; no wall-clock waits.
Fabricated data only.
"""

from __future__ import annotations

import io
import os
from typing import IO, TYPE_CHECKING

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.gui.exporters.csv_exporter import CsvExporter
from src.gui.exporters.excel_exporter import ExcelExporter
from src.gui.exporters.registry import ExporterRegistry
from src.gui.pipeline_service import ImportSpec, PipelineService
from src.gui.presenters.export_presenter import ExportPresenter
from src.gui.presenters.pipeline_presenter import PipelinePresenter
from src.gui.services.db_service import DbService
from src.pandas_io import read_excel_sheet
from tests.aop_fixtures import aop_header_without_key, make_aop_row
from tests.gui.fakes.fake_views import FakeExportView, FakePipelineView
from tests.le_fixtures import make_row, patch_connect, source_header_without_key

if TYPE_CHECKING:
    from collections.abc import Callable

    import pytest

# Fabricated GtN labels and AOP type rows shared across LE and AOP sheets so the
# pipeline produces normal, tie-able lines.
_GTN_LABELS = ["Gross Sales", "Lbs", "Off Invoice", "Trade", "Non-Trade"]
_AOP_TYPES = [
    ("Gross Sales", 100.0),
    ("LBs", 10.0),
    ("Off Invoice $", 5.0),
    ("Trade Spend $", 2.0),
    ("Non-Trade $", 1.0),
]
_SKUS = ("SKU-001", "SKU-002")


def _le_rows() -> list[dict[str, object]]:
    """Fabricated LE source rows: one GtN row per (SKU, label)."""
    rows: list[dict[str, object]] = []
    for sku in _SKUS:
        for label in _GTN_LABELS:
            value = 12.0 if label == "Lbs" else 100.0
            months = [0.0, 0.0, 0.0, 0.0, *([value / 8.0] * 8)]
            rows.append(
                make_row(
                    customer="Acme Foods",
                    sku=sku,
                    type_=label,
                    ppg="PPG-1",
                    months=months,
                    gtn=label,
                    description="Widget",
                )
            )
    return rows


def _aop_rows() -> list[dict[str, object]]:
    """Fabricated AOP source rows: one per (SKU, type)."""
    rows: list[dict[str, object]] = []
    for sku in _SKUS:
        for type_label, ytg in _AOP_TYPES:
            months = [0.0, 0.0, 0.0, 0.0, *([ytg / 8.0] * 8)]
            rows.append(
                make_aop_row(
                    customer="Acme Foods",
                    sku=sku,
                    type_=type_label,
                    months=months,
                    customer_master="Master Group",
                    description="Widget",
                    ppg="PPG-1",
                )
            )
    return rows


def _build_combined_workbook() -> io.BytesIO:
    """Build a single workbook carrying LE, AOP, and SKU_LU sheets in memory."""
    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)

    # LE sheet (header on Excel row 3).
    le_sheet = workbook.create_sheet("LE-8 + 4")
    le_header = source_header_without_key()
    le_sheet.append(["leading note row 1"])
    le_sheet.append(["leading note row 2"])
    le_sheet.append(le_header)
    for record in _le_rows():
        le_sheet.append([record.get(column) for column in le_header])

    # AOP sheet (header on Excel row 3).
    aop_sheet = workbook.create_sheet("AOP1")
    aop_header = aop_header_without_key()
    aop_sheet.append(["leading note row 1"])
    aop_sheet.append(["leading note row 2"])
    aop_sheet.append(aop_header)
    for record in _aop_rows():
        aop_sheet.append([record.get(column) for column in aop_header])

    # SKU_LU sheet (header on Excel row 1) — fabricated values only.
    sku_sheet = workbook.create_sheet("SKU_LU")
    sku_sheet.append(["SKU", "SKU Description", "Category", "International"])
    sku_sheet.append(["SKU-001", "Widget A", "Category X", 0])
    sku_sheet.append(["SKU-002", "Widget B", "Category Y", 1])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _patch_loaders(monkeypatch: pytest.MonkeyPatch, buffer: io.BytesIO) -> None:
    """Patch the three loader reads to read a single shared in-memory buffer."""
    from src import load_aop, load_skulu, normalize_le

    real_load_source = normalize_le.load_source
    real_load_aop = load_aop.load_aop
    real_load_skulu = load_skulu.load_skulu

    def _fake_load_source(
        _path: str, sheet: str, *, key_mismatch: str = "prompt"
    ) -> object:
        buffer.seek(0)
        return real_load_source(buffer, sheet, key_mismatch=key_mismatch)

    def _fake_load_aop(
        _path: str, *, sheet: str = "AOP1", key_mismatch: str = "prompt"
    ) -> object:
        buffer.seek(0)
        return real_load_aop(buffer, sheet=sheet, key_mismatch=key_mismatch)

    def _fake_load_skulu(_path: str, *, sheet: str = "SKU_LU") -> object:
        buffer.seek(0)
        return real_load_skulu(buffer, sheet=sheet)

    monkeypatch.setattr("src.normalize_le.load_source", _fake_load_source)
    monkeypatch.setattr("src.load_aop.load_aop", _fake_load_aop)
    monkeypatch.setattr("src.load_skulu.load_skulu", _fake_load_skulu)


class _NonClosingStringIO(io.StringIO):
    """A StringIO whose close is a no-op so captured CSV text survives the with."""

    def close(self) -> None:
        """Suppress closing so the captured text persists for assertions."""
        return None


def _capturing_csv_writer() -> (
    tuple[dict[str, _NonClosingStringIO], Callable[[str], IO[str]]]
):
    """Return a (captures, open_writer) pair for the CSV exporter."""
    captures: dict[str, _NonClosingStringIO] = {}

    def _open_writer(path: str) -> IO[str]:
        sink = _NonClosingStringIO()
        captures[path] = sink
        return sink

    return captures, _open_writer


def test_gui_end_to_end_pipeline_round_trip(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Drive select -> import -> run -> save -> reopen -> export end-to-end.

    Uses the real PipelineService, DbService, and exporters with fakes only at
    the Qt-view boundary. Verifies the derived-table set survives a save/reopen
    and that both exports contain the selected tables.
    """
    # Arrange: in-memory SQLite, in-memory workbooks, real service stack.
    con = patch_connect(monkeypatch)
    buffer = _build_combined_workbook()
    _patch_loaders(monkeypatch, buffer)

    service = PipelineService(db_service=DbService())
    pipeline_view = FakePipelineView()
    pipeline_presenter = PipelinePresenter(pipeline_view, service)

    spec = ImportSpec(
        le_path="combined.xlsx",
        le_sheet="LE-8 + 4",
        aop_path="combined.xlsx",
        aop_sheet="AOP1",
        skulu_path="combined.xlsx",
        skulu_sheet="SKU_LU",
    )

    # Act 1 — import all and run the pipeline through the presenter.
    pipeline_presenter.on_import_all(spec)
    pipeline_presenter.on_run()

    # Assert 1 — derived tables produced and success reported.
    derived_keys_before = set(pipeline_presenter.derived_tables)
    assert "mix_rollup_4" in derived_keys_before
    assert any("Run complete" in s for s in pipeline_view.results)

    # Act 2 — save to an in-memory SQLite, then reopen into the working set.
    pipeline_presenter.on_save("results.db")
    assert pipeline_presenter.db_path == "results.db"
    pipeline_presenter.on_open_db("results.db")

    # Assert 2 — the working set after reopen carries the derived keys.
    assert derived_keys_before.issubset(set(pipeline_presenter.imported_tables))

    # Act 3 — set up the export registry and presenter; export to Excel BytesIO.
    registry = ExporterRegistry()
    excel_target = io.BytesIO()
    captures, csv_writer = _capturing_csv_writer()
    registry.register(ExcelExporter())
    registry.register(CsvExporter(open_writer=csv_writer))

    export_view = FakeExportView(selected=["mix_rollup_4"])
    export_presenter = ExportPresenter(export_view, registry)
    tables = pipeline_presenter.imported_tables

    # Excel export to the in-memory buffer. Use a concrete ExcelExporter
    # directly because its destination supports IO[bytes] while the Protocol
    # narrows to str for the production path-based UX. The presenter -> exporter
    # routing for the path case is already covered in test_export_presenter.
    excel_exporter = ExcelExporter()
    excel_exporter.export(tables, export_view.get_selected_names(), excel_target)

    # CSV export to a fabricated destination path; the injected writer captures
    # content. v2 Decision 1: the destination is now a CSV file path, not a
    # directory; the exporter writes one file per selected table using the
    # ``<base>_<table>.csv`` name-mangling rule (here base == "results").
    export_presenter.on_export(tables, "CSV", "out-dir/results.csv")

    # Assert 3 — both exports contain the selected table.
    excel_target.seek(0)
    workbook = load_workbook(excel_target, read_only=True)
    try:
        assert "mix_rollup_4" in workbook.sheetnames
    finally:
        workbook.close()
    excel_target.seek(0)
    sheet_back = read_excel_sheet(excel_target, sheet_name="mix_rollup_4", header=0)
    assert "value" in sheet_back.columns

    # v2 name-mangled path: "<directory>/<base>_<table>.csv".
    csv_path = os.path.join("out-dir", "results_mix_rollup_4.csv")
    assert csv_path in captures
    csv_back = pd.read_csv(io.StringIO(captures[csv_path].getvalue()))
    assert list(csv_back.columns) == ["value"]

    con.real_close()
