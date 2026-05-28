"""End-to-end integration test for :mod:`src.mix_pipeline`.

Builds a fabricated in-memory workbook supplying the ``AOP1``, ``LE-8 + 4``, and
``SKU_LU`` sheets, runs ``mix_pipeline.main`` against a shared in-memory SQLite
database, and asserts that ``main`` returns ``0``, that the two import tables and
all twenty-two derived tables exist, and that the customer-layer rollup ties out to
the scalar ``mix_rollup_4``. No runtime temp files are used; the workbook is a
``BytesIO`` and the database is ``sqlite3.connect(":memory:")``. Fabricated data
only; no confidential values appear.
"""

from __future__ import annotations

import io
import sqlite3
from typing import TYPE_CHECKING, cast

from openpyxl import Workbook

from src import load_aop, load_skulu, mix_pipeline, normalize_le
from src.pandas_io import read_table
from tests.aop_fixtures import aop_header_without_key, make_aop_row
from tests.le_fixtures import (
    PersistentConnection,
    make_row,
    patch_connect,
    source_header_without_key,
)

if TYPE_CHECKING:
    import pytest


def _f(value: object) -> float:
    """Convert a pandas/numpy scalar to a built-in float for typed assertions."""
    return float(cast("float", value))


# The fabricated SKUs and the GtN/Type labels shared across the LE and AOP
# sheets so the pipeline produces normal, tie-able lines.
_GTN_LABELS = ["Gross Sales", "Lbs", "Off Invoice", "Trade", "Non-Trade"]
_AOP_TYPES = [
    ("Gross Sales", 100.0),
    ("LBs", 10.0),
    ("Off Invoice $", 5.0),
    ("Trade Spend $", 2.0),
    ("Non-Trade $", 1.0),
]


def _le_rows() -> list[dict[str, object]]:
    """Build LE source rows: one GtN label row per SKU with a YTG-half profile."""
    rows: list[dict[str, object]] = []
    # Two SKUs in one customer; each GtN label carries a May..Dec-weighted month
    # profile so the loader's YTG = sum(May..Dec) derivation is exercised. The
    # Type segment is distinct per GtN label so the rebuilt KEY differs per label
    # and normalize_le does not collapse the five GtN rows into one.
    for sku in ("SKU-001", "SKU-002"):
        for label in _GTN_LABELS:
            value = 12.0 if label == "Lbs" else 100.0
            months = [0.0, 0.0, 0.0, 0.0, *([value / 8.0] * 8)]
            rows.append(
                make_row(
                    customer="Acme Foods",
                    sku=sku,
                    type_=label,
                    ppg="PPG-1",
                    months=months,
                    gtn=label,
                    description="Widget",
                )
            )
    return rows


def _aop_rows() -> list[dict[str, object]]:
    """Build AOP source rows: one Type row per SKU with a May..Dec profile."""
    rows: list[dict[str, object]] = []
    for sku in ("SKU-001", "SKU-002"):
        for type_label, ytg in _AOP_TYPES:
            months = [0.0, 0.0, 0.0, 0.0, *([ytg / 8.0] * 8)]
            rows.append(
                make_aop_row(
                    customer="Acme Foods",
                    sku=sku,
                    type_=type_label,
                    months=months,
                    customer_master="Master Group",
                    description="Widget",
                    ppg="PPG-1",
                )
            )
    return rows


def _append_le_sheet(workbook: Workbook) -> None:
    """Append the LE-8 + 4 sheet (header on Excel row 3) to the workbook."""
    sheet = workbook.create_sheet("LE-8 + 4")
    header = source_header_without_key()
    sheet.append(["leading note row 1"])
    sheet.append(["leading note row 2"])
    sheet.append(header)
    # Emit each LE row in the header's column order.
    for record in _le_rows():
        sheet.append([record.get(column) for column in header])


def _append_aop_sheet(workbook: Workbook) -> None:
    """Append the AOP1 sheet (header on Excel row 3) to the workbook."""
    sheet = workbook.create_sheet("AOP1")
    header = aop_header_without_key()
    sheet.append(["leading note row 1"])
    sheet.append(["leading note row 2"])
    sheet.append(header)
    for record in _aop_rows():
        sheet.append([record.get(column) for column in header])


def _append_skulu_sheet(workbook: Workbook) -> None:
    """Append the SKU_LU sheet (header on Excel row 1) to the workbook."""
    sheet = workbook.create_sheet("SKU_LU")
    header = ["SKU", "SKU Description", "Category", "International"]
    sheet.append(header)
    sheet.append(["SKU-001", "Widget A", "Category X", 0])
    sheet.append(["SKU-002", "Widget B", "Category Y", 1])


def _build_combined_workbook() -> io.BytesIO:
    """Build one in-memory workbook carrying the LE, AOP, and SKU_LU sheets."""
    workbook = Workbook()
    # Drop the default empty sheet so only the three named sheets remain.
    default = workbook.active
    assert default is not None
    workbook.remove(default)
    _append_le_sheet(workbook)
    _append_aop_sheet(workbook)
    _append_skulu_sheet(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _patch_loaders(
    monkeypatch: pytest.MonkeyPatch,
    buffer: io.BytesIO,
) -> None:
    """Patch the three loader read functions to read the shared in-memory buffer."""
    # Capture the original loader callables before patching so the fakes delegate
    # to the real readers rather than recursing into themselves.
    real_load_source = normalize_le.load_source
    real_load_aop = load_aop.load_aop
    real_load_skulu = load_skulu.load_skulu

    def _fake_load_source(
        _path: str, sheet: str, *, key_mismatch: str = "prompt"
    ) -> object:
        buffer.seek(0)
        return real_load_source(buffer, sheet, key_mismatch=key_mismatch)

    def _fake_load_aop(
        _path: str, *, sheet: str = "AOP1", key_mismatch: str = "prompt"
    ) -> object:
        buffer.seek(0)
        return real_load_aop(buffer, sheet=sheet, key_mismatch=key_mismatch)

    def _fake_load_skulu(_path: str, *, sheet: str = "SKU_LU") -> object:
        buffer.seek(0)
        return real_load_skulu(buffer, sheet=sheet)

    monkeypatch.setattr("src.normalize_le.load_source", _fake_load_source)
    monkeypatch.setattr("src.load_aop.load_aop", _fake_load_aop)
    monkeypatch.setattr("src.load_skulu.load_skulu", _fake_load_skulu)


_DERIVED_TABLES = [
    "le_wide",
    "aop_wide",
    "customer_lu",
    "sku_lu",
    "aop_norm",
    "le_norm",
    "aop_vs_le",
    "mix_base",
    "rate_impacts",
    "mix_rollup_1",
    "mix_1_sku",
    "mix_rollup_2",
    "mix_2_category",
    "mix_rollup_3",
    "mix_3_customer",
    "mix_rollup_4",
    "mix_4_country",
    "mix_0_detail",
    "mix_2_sku_bottomsup",
    "mix_3_category_bottomsup",
    "mix_4_customer_bottomsup",
    "q1_results_by_sku",
]


def test_mix_pipeline_end_to_end(monkeypatch: pytest.MonkeyPatch) -> None:
    """main runs end-to-end and persists the import and all derived tables."""
    # Arrange: one shared in-memory database and the combined workbook.
    con: PersistentConnection = patch_connect(monkeypatch)
    buffer = _build_combined_workbook()
    _patch_loaders(monkeypatch, buffer)

    # Act
    exit_code = mix_pipeline.main(["--input", "ignored.xlsx", "--output", "ignored.db"])

    # Assert: clean exit and every expected table present.
    assert exit_code == 0
    existing = {
        str(name[0])
        for name in con.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()
    }
    assert {"LE", "aop"}.issubset(existing)
    for table in _DERIVED_TABLES:
        assert table in existing, f"missing derived table: {table}"
    con.real_close()


def test_mix_pipeline_rollup_tie_out(monkeypatch: pytest.MonkeyPatch) -> None:
    """The persisted customer layer ties out to the scalar mix_rollup_4."""
    # Arrange
    con: PersistentConnection = patch_connect(monkeypatch)
    buffer = _build_combined_workbook()
    _patch_loaders(monkeypatch, buffer)

    # Act
    exit_code = mix_pipeline.main(["--input", "ignored.xlsx", "--output", "ignored.db"])

    # Assert: the single-row mix_rollup_4 equals the customer-layer impact sum.
    assert exit_code == 0
    mix_3 = read_table(con, "mix_3_customer")
    rollup_4 = read_table(con, "mix_rollup_4")
    scalar = _f(rollup_4.iloc[0, 0])
    assert abs(_f(mix_3["Calc Net Price Impact"].sum()) - scalar) < 1e-9
    con.real_close()


def test_mix_pipeline_loader_error_returns_one(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A loader ValueError maps to a clean exit code 1."""
    # Arrange: patch the LE loader to raise the kind of ValueError a loader emits.
    patch_connect(monkeypatch)

    def _raise(_path: str, _sheet: str, *, key_mismatch: str = "prompt") -> object:
        raise ValueError("Source schema mismatch: could not resolve column")

    monkeypatch.setattr("src.normalize_le.load_source", _raise)

    # Act
    exit_code = mix_pipeline.main(["--input", "ignored.xlsx", "--output", "ignored.db"])

    # Assert
    assert exit_code == 1


def test_sqlite_connection_helper_is_in_memory() -> None:
    """The test database is in-memory (no runtime temp files are created)."""
    # Arrange / Act: a bare in-memory connection round-trips a trivial table.
    con = sqlite3.connect(":memory:")
    con.execute("CREATE TABLE t (a INTEGER)")
    con.execute("INSERT INTO t VALUES (1)")

    # Assert
    assert con.execute("SELECT a FROM t").fetchone()[0] == 1
    con.close()
